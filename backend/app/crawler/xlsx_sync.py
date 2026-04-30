#!/usr/bin/env python3
"""
FDA XLSX 同步脚本
下载 FDA 官方 Warning Letters 数据包，覆盖/增量更新数据库
"""
import os
import sys
import sqlite3
import tempfile
from datetime import datetime, date

import httpx
import openpyxl  # pip install openpyxl

DB_PATH = '/root/data/fda_warning.db'
XLSX_URL = (
    "https://www.fda.gov/inspections-compliance-enforcement-and-criminal-investigations/"
    "compliance-actions-and-activities/warning-letters/datatables-data?page&_format=xlsx"
)
HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; FDA-WarningLetter-Bot/1.0; +https://19990419.top)"
}


def parse_date(val) -> date | None:
    """把 Excel 日期值转成 date，或从字符串 parse"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            return openpyxl.cell.Cell().value_to_maybe_date(val)
        except Exception:
            return None
    s = str(val).strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None


def download_xlsx() -> str:
    """下载 XLSX 到临时文件，返回路径"""
    print(f"📥 下载 XLSX from FDA...")
    with httpx.Client(timeout=60.0, headers=HEADERS, follow_redirects=True) as client:
        resp = client.get(XLSX_URL)
        resp.raise_for_status()
    suffix = ".xlsx"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as f:
        f.write(resp.content)
        path = f.name
    print(f"✅ 下载完成: {len(resp.content):,} bytes → {path}")
    return path


def sync_xlsx(xlsx_path: str):
    """解析 XLSX，对比数据库，增量更新"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # 读取 header
    headers = [str(c.value).strip() if c.value else f"col_{i}" for i, c in enumerate(ws[1])]
    print(f"表头 ({len(headers)} 列): {headers[:8]}...")

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    updated = 0
    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row):
            continue

        # 构建 dict（按表头索引）
        rec = dict(zip(headers, row))

        fda_id = str(rec.get("FEI Number") or rec.get("FEI_Number") or rec.get("FEI") or "").strip()
        company = str(rec.get("Company Name") or rec.get("Company") or "").strip()
        subject = str(rec.get("Subject") or rec.get("Subject / Issue") or "").strip()
        office = str(rec.get("Issuing Office") or rec.get("Office") or "").strip()
        issue_date = parse_date(rec.get("Issue Date") or rec.get("Date Issued"))
        posted_date = parse_date(rec.get("Posted Date") or rec.get("Post Date"))
        closeout = parse_date(rec.get("Closeout Date") or rec.get("Closeout"))
        response = parse_date(rec.get("Response Date") or rec.get("Response"))
        status_val = str(rec.get("Status") or "Open").strip()
        country = str(rec.get("Country") or rec.get("Location") or "").strip()
        closeout_pdf = str(rec.get("Closeout Letter URL") or rec.get("Closeout PDF") or "").strip()
        slug_val = str(rec.get("URL Slug") or rec.get("Slug") or "").strip()

        if not fda_id and not company:
            skipped += 1
            continue

        # 检查是否已存在
        cursor.execute(
            "SELECT id, url FROM warning_letters WHERE fda_id = ? OR company_name = ? LIMIT 1",
            (fda_id, company)
        )
        existing = cursor.fetchone()

        if existing:
            # 更新已有记录
            cursor.execute("""
                UPDATE warning_letters SET
                    company_name = COALESCE(NULLIF(?, ''), company_name),
                    subject = COALESCE(NULLIF(?, ''), subject),
                    issuing_office = COALESCE(NULLIF(?, ''), issuing_office),
                    issue_date = COALESCE(?, issue_date),
                    posted_date = COALESCE(?, posted_date),
                    closeout_date = COALESCE(?, closeout_date),
                    response_date = COALESCE(?, response_date),
                    status = COALESCE(NULLIF(?, ''), status),
                    country = COALESCE(NULLIF(?, ''), country),
                    closeout_pdf_url = COALESCE(NULLIF(?, ''), closeout_pdf_url),
                    slug = COALESCE(NULLIF(?, ''), slug),
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (company, subject, office, issue_date, posted_date,
                  closeout, response, status_val, country, closeout_pdf,
                  slug_val, existing[0]))
            updated += 1
        else:
            # 插入新记录
            cursor.execute("""
                INSERT INTO warning_letters
                    (fda_id, company_name, subject, issuing_office, issue_date,
                     posted_date, closeout_date, response_date, status, country,
                     closeout_pdf_url, slug, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
            """, (fda_id, company, subject, office, issue_date,
                  posted_date, closeout, response, status_val, country,
                  closeout_pdf, slug_val))
            inserted += 1

    conn.commit()

    cursor.execute("SELECT COUNT(*) FROM warning_letters")
    total = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM warning_letters WHERE url IS NOT NULL AND url != ''")
    with_url = cursor.fetchone()[0]

    print(f"\n📊 同步结果:")
    print(f"   新增: {inserted}")
    print(f"   更新: {updated}")
    print(f"   跳过: {skipped}")
    print(f"   总计: {total} 封信, {with_url} 条有URL")

    conn.close()
    os.unlink(xlsx_path)


def main():
    try:
        path = download_xlsx()
        sync_xlsx(path)
    except Exception as e:
        print(f"❌ XLSX 同步失败: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
